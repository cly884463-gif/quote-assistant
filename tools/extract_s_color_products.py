import json
from pathlib import Path

import openpyxl


SOURCE = Path(r"F:\色浆配比清单-2026.xlsx")
OUTPUT = Path("utils/s-color-products.js")


def clean(value):
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def is_marked(row_cells):
    return any(cell.fill.patternType and cell.fill.patternType != "none" for cell in row_cells)


def build_product():
    ws = openpyxl.load_workbook(SOURCE, data_only=True).active
    spec_options = []

    for row in range(2, ws.max_row + 1):
        cells = [ws.cell(row, col) for col in range(1, 8)]
        values = [cell.value for cell in cells]
        if not values[1] or not values[2]:
            continue

        paint_type = "清漆" if is_marked(cells) else "白漆"
        code = str(values[1])
        name = str(values[2])
        package_spec = f"{clean(values[3])}ml/瓶"
        price = clean(values[4])
        ratio_gg = str(values[5] or "")
        ratio_gml = str(values[6] or "")

        spec_options.append({
            "spec": f"{code} {name}（{paint_type}）{package_spec}",
            "workTimes": "",
            "coverage": f"{clean(values[3])}ml",
            "unit": "瓶",
            "packageSpec": package_spec,
            "dealerPrice": price,
            "costPerSquare": "",
            "remark": f"漆种：{paint_type}；产品编号：{code}；产品名称：{name}；配比(g:g)：{ratio_gg}；配比(g:ml)：{ratio_gml}"
        })

    return {
        "id": "gn-s-sj-高浓度s系列色浆",
        "model": "GN-S-SJ",
        "category": "高浓度S系列色浆",
        "name": "高浓度S系列色浆",
        "specs": [option["spec"] for option in spec_options],
        "specOptions": spec_options,
        "workTimes": "",
        "coverage": "200ml/300ml",
        "unit": "瓶",
        "packageSpec": "200ml/瓶、300ml/瓶",
        "dealerPrice": spec_options[0]["dealerPrice"] if spec_options else "",
        "costPerSquare": "",
        "remark": "高浓度S系列色浆，规格中已标注清漆/白漆，备注含配比详情"
    }


def main():
    product = build_product()
    content = (
        "// Auto-generated from F:/色浆配比清单-2026.xlsx\n"
        "// Colored rows are treated as 清漆; uncolored rows are treated as 白漆.\n"
        f"const sColorPasteProduct = {json.dumps(product, ensure_ascii=False, indent=2)};\n\n"
        "module.exports = {\n"
        "  sColorPasteProduct\n"
        "};\n"
    )
    OUTPUT.write_text(content, encoding="utf-8")
    print(len(product["specOptions"]))


if __name__ == "__main__":
    main()
