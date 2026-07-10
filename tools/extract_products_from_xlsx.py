import argparse
import json
import re
from pathlib import Path

import openpyxl


DEFAULT_SOURCE = r"F:\报价体系及合同\新材联-经销商批发报价最新.xlsx"
DEFAULT_OUTPUT = Path("utils/products.js")


def clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def make_id(value):
    return re.sub(r"[^a-zA-Z0-9\u4e00-\u9fff]+", "-", value).strip("-").lower()


def round_price(value):
    if isinstance(value, (int, float)):
        return int(value + 0.5)
    return value


def calculate_cost_per_square(price, coverage, fallback):
    if isinstance(price, (int, float)) and isinstance(coverage, (int, float)) and coverage:
        return price / coverage
    return clean(fallback)


def extract_products(source, should_round_prices=False):
    ws = openpyxl.load_workbook(source, data_only=True).active
    grouped = {}
    current = {"model": "", "category": "", "name": ""}

    for row_index in range(7, ws.max_row + 1):
        values = [ws.cell(row_index, col).value for col in range(1, 13)]
        if values[0] and str(values[0]).startswith("合计"):
            break

        row_has_new_identity = values[1] not in (None, "") or values[2] not in (None, "")
        if values[0] not in (None, ""):
            current["model"] = str(values[0])
        elif row_has_new_identity and values[2] not in (None, ""):
            current["model"] = "未编号"

        if values[1] not in (None, ""):
            current["category"] = str(values[1])
        if values[2] not in (None, ""):
            current["name"] = str(values[2])

        spec = values[3]
        price = values[8]
        if not current["name"] or spec in (None, "") or price in (None, ""):
            continue
        normalized_price = round_price(price) if should_round_prices else clean(price)
        cost_per_square = calculate_cost_per_square(normalized_price, values[5], values[10])

        model = current["model"] or "未编号"
        group_key = f"{model}|{current['category']}|{current['name']}"
        if group_key not in grouped:
            grouped[group_key] = {
                "id": make_id(f"{model}-{current['name']}"),
                "model": model,
                "category": current["category"],
                "name": current["name"],
                "specs": [],
                "specOptions": [],
                "workTimes": clean(values[4]),
                "coverage": clean(values[5]),
                "unit": clean(values[6]),
                "dealerPrice": normalized_price,
                "costPerSquare": cost_per_square,
                "remark": clean(values[11]),
            }

        product = grouped[group_key]
        spec_text = str(spec)
        option = {
            "spec": spec_text,
            "workTimes": clean(values[4]),
            "coverage": clean(values[5]),
            "unit": clean(values[6]),
            "dealerPrice": normalized_price,
            "costPerSquare": cost_per_square,
            "remark": clean(values[11]),
        }
        product["specs"].append(spec_text)
        product["specOptions"].append(option)

    return list(grouped.values())


def build_content(products, source):
    source_label = str(source).replace("\\", "/")
    return (
        f"// Auto-generated from {source_label}\n"
        "// Same product cards are grouped by model/category/name; specOptions keep per-spec prices accurate.\n"
        f"const products = {json.dumps(products, ensure_ascii=False, indent=2)};\n\n"
        "function filterProducts(keyword) {\n"
        "  const normalized = String(keyword || \"\").trim().toLowerCase();\n"
        "  if (!normalized) {\n"
        "    return products;\n"
        "  }\n\n"
        "  return products.filter((product) => {\n"
        "    const text = [\n"
        "      product.model,\n"
        "      product.category,\n"
        "      product.name,\n"
        "      product.specs.join(\" \")\n"
        "    ].join(\" \").toLowerCase();\n"
        "    return text.includes(normalized);\n"
        "  });\n"
        "}\n\n"
        "function groupProductsByCategory(sourceProducts) {\n"
        "  const groups = [];\n"
        "  const groupMap = {};\n"
        "  (sourceProducts || products).forEach((product) => {\n"
        "    const category = product.category || \"未分类\";\n"
        "    if (!groupMap[category]) {\n"
        "      groupMap[category] = {\n"
        "        category,\n"
        "        open: groups.length === 0,\n"
        "        products: []\n"
        "      };\n"
        "      groups.push(groupMap[category]);\n"
        "    }\n"
        "    groupMap[category].products.push(product);\n"
        "  });\n"
        "  return groups;\n"
        "}\n\n"
        "module.exports = {\n"
        "  products,\n"
        "  filterProducts,\n"
        "  groupProductsByCategory\n"
        "};\n"
    )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", default=DEFAULT_SOURCE)
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--round-prices", action="store_true")
    args = parser.parse_args()

    source = Path(args.source)
    output = Path(args.output)
    products = extract_products(source, args.round_prices)
    output.write_text(build_content(products, source), encoding="utf-8")
    print(len(products))


if __name__ == "__main__":
    main()
